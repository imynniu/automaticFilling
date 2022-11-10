import openpyxl
import requests
from time import sleep
import time
import json

# 请求头
my_header = {
    "User-Agent":"User-Agent: Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.143 Safari/537.36 MicroMessenger/7.0.9.501 NetType/WIFI MiniProgramEnv/Windows WindowsWechat",
    "Referer":"https://servicewechat.com/wx23d8d7ea22039466/1674/page-frame.html",
}
# "imprint": "oWRkU0X0y2TYFRDqFFdGW153oLM0",
# "Accept-Encoding": "gzip, deflate, br",
# "Connection": " keep-alive",
# "Host": "a.welife001.com"
##url 地址
host = 'https://a.welife001.com'


##参数定义
# 提交地址
url_submit ='/applet/notify/feedbackWithOss'

# 数据源
filename = 'data.xlsx'
sheet_name = 'Sheet1'
filename_data = ''
file_log = 'log.txt'

# 日志输出方式
# with open("text.txt","a") as file:
#     file.write("What I want to add on goes here")

# 读取数据
data = []
dict_data = {}

# 读取excel数据
def read_data(filename,sheet_name):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=sheet.max_column, max_row=sheet.max_row):
        temp_data = []
        for d in row :
            temp_data.append(d.value)
        data.append(temp_data)
    return 1
#读取请求体数据
# def read_request_body(filename):
#     with open(filename,'r',encoding="utf-8") as file:
#         return file.read()
        ############
        # print(file.read())
        # dict_body =  file.read()
        # print(dict_body)
    # return dict_body
#读取请求体数据
def read_request_body(filename):
    #修改请求头
    global dict_data 
    global my_header
    dict_data = {}
    my_header = {}
    with open(filename, "r",encoding='utf-8') as f:
        for line in f.readlines():
            line = line.strip('\n')  #去掉列表中每一个元素的换行符
            # print(line.split(':'))
            list_line = line.split(':')
            if list_line[0] == 'User-Agent':
                my_header['User-Agent'] = list_line[1].strip()
            elif list_line[0] == 'imprint':
                my_header['imprint'] = list_line[1].strip()
            elif list_line[0] == 'Referer':
                # print(list_line)
                my_header['Referer'] = list_line[1].strip()+':'+list_line[2]
    #修改请求体
    with open(filename, 'r',encoding='utf-8') as fp:
        lines = fp.readlines()
        dict_data = eval(lines[-1])
#写日志
def write_log(file_log,content):
    with open(file_log,"a") as file:
        file.write(content)
# 修改请求头
def modify_header(data_person):
    my_header['imprint'] = data_person[12]
# 修改信息
def modify_data(data_person):
    # 问卷id
    # id_person = data_person[0]
    # dict_data['id'] = id_person
    # 打卡日期
    time_today = time.strftime('%Y-%m-%d', time.localtime(time.time()))
    dict_data['daka_day'] = time_today
    # # 成员id
    member_id_person = data_person[1]
    dict_data['member_id'] = member_id_person

    # 问卷日期
    dict_data['invest']['day'] = time_today
    # 打卡时间戳
    dict_data['invest']['time'] = int(time.time()*1000)
    # 体温
    #体温
    temperature_person =data_person[0]
    dict_data['invest']['subject'][1]['input']['content']=temperature_person
    # 地理位置
    city = data_person[3]
    buildings = data_person[4]
    lat = data_person[5]
    lng = data_person[6]
    loc = {
        "title":buildings,
        "address":city,
        "location":[float(lat), float(lng)]
    }
    dict_data['invest']['subject'][2]['input']['content'] = json.dumps(loc)
    # 联系赵洪吉
    content_person = data_person[7]
    dict_data['invest']['subject'][5]['input']['content'] = content_person




###提交表单
def send_request(url,request_data,log_name):
    r= requests.post(url,json=request_data, headers= my_header)
    if(r.json()['msg'] =='ok'):
        write_log(file_log,time.strftime('\n'+'%Y-%m-%d %H:%M:%S', time.localtime(time.time()))+' '+log_name+' 的信息上传成功:)')
        # with open("log.txt","a") as file:
        #     file.write(time.strftime('\n'+'%Y-%m-%d %H:%M:%S', time.localtime(time.time()))+' '+log_name+' 的信息上传成功！！')
    else:
        write_log(file_log, time.strftime('\n' + '%Y-%m-%d %H:%M:%S',
                                          time.localtime(time.time())) + ' ' + log_name + ' 的信息上传失败:(')
    return 1


if __name__ == '__main__':
    while(1):
        data.clear()
        re_read = read_data(filename, sheet_name)
        for data_temp in data:
            dict_data = {}
            my_header = {}
            filename_data = data_temp[9]
            read_request_body(filename_data)
            modify_data(data_temp)
            url =host+url_submit
            #测试查看数据、
            # print(dict_data)
            # print(my_header)
            # print(url)
            
            send_request(url,dict_data,str(data_temp[8]))
            sleep(5)
        print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))
        print('填报成功，等待八小时后再次填报......')
        sleep(28800)